# evaluator.py

import os
import sys
import time
import pandas as pd
from typing import List, Tuple, Optional
from openai import OpenAI
from dotenv import load_dotenv

from constants import INPUT_EXCEL_FILENAME, OUTPUT_EXCEL_FILENAME, OUTPUT_EVALUATION_FILENAME


# Загружаем переменные окружения из .env
load_dotenv()
API_KEY = os.getenv("BOTHUB_API_KEY")

API_TIMEOUT = 5

# 🔍 Проверка API-ключа

if not API_KEY:
    print("[CRITICAL] ❌ Переменная BOTHUB_API_KEY не установлена!")
    print("👉 Проверь, что в корне проекта есть файл .env с содержимым:")
    print("   BOTHUB_API_KEY=твой_реальный_токен_сюда")
    sys.exit(1)

if API_KEY == "<your bothub access token>" or API_KEY.strip() == "":
    print("[CRITICAL] ❌ BOTHUB_API_KEY содержит placeholder или пуст!")
    print("👉 Замени '<your bothub access token>' в коде или, лучше, задай ключ в .env")
    sys.exit(1)

if len(API_KEY) < 10:
    print("[WARNING] ⚠️ BOTHUB_API_KEY подозрительно короткий. Проверь, что ключ полный.")

print(f"[INFO] ✅ BOTHUB_API_KEY успешно загружен (первые 5 символов: {API_KEY[:5]}...)")



# Инициализация клиента Bothub
client = OpenAI(
    api_key=API_KEY,
    base_url="https://bothub.chat/api/v2/openai/v1"
)

MODEL_NAME = "deepseek-chat-v3-0324:free"


def load_data() -> List[Tuple[int, str, str]]:
    """
    Загружает эталонные ответы и ответы модели, сопоставляет по порядку.

    Столбцы:
        questions.xlsx: столбец D — эталонный ответ
        answers_v1.xlsx: столбец "Ответ модели" — ответ локальной модели

    Возвращает:
        List[Tuple[int, str, str]]: [(номер, эталон, ответ_модели), ...]
    """
    try:
        # Загружаем эталонные ответы (столбец D = 3-й индекс)
        df_questions = pd.read_excel(INPUT_EXCEL_FILENAME, header=0)
        if df_questions.shape[1] < 4:
            raise ValueError("В questions.xlsx меньше 4 столбцов. Ожидался столбец D с эталонными ответами.")
        ground_truths = df_questions.iloc[:, 3].fillna("").astype(str).tolist()  # Столбец D (индекс 3)

        # Загружаем ответы модели
        df_answers = pd.read_excel(OUTPUT_EXCEL_FILENAME, header=0)
        if "Ответ модели" not in df_answers.columns:
            raise ValueError("В answers_v1.xlsx нет столбца 'Ответ модели'")
        model_answers = df_answers["Ответ модели"].fillna("").astype(str).tolist()

        # Убедимся, что списки одинаковой длины
        min_len = min(len(ground_truths), len(model_answers))
        if len(ground_truths) != len(model_answers):
            print(f"[WARNING] Длина ответов не совпадает: эталонов {len(ground_truths)}, ответов модели {len(model_answers)}. Будет обработано {min_len} записей.")

        # Формируем список кортежей
        data = []
        for i in range(min_len):
            data.append((i + 1, ground_truths[i], model_answers[i]))

        return data

    except FileNotFoundError as e:
        print(f"[ERROR] Файл не найден: {e}")
        return []
    except Exception as e:
        print(f"[ERROR] Ошибка загрузки данных: {e}")
        return []


def ask_llm_for_similarity(ground_truth: str, model_answer: str, model: str = MODEL_NAME, max_retries: int = 3) -> Optional[float]:
    """
    Отправляет два текста в LLM и получает оценку соответствия от 0 до 1.

    Возвращает:
        float: оценка от 0.0 до 1.0, или None в случае ошибки.
    """
    prompt = (
        f"Вот два текста.\n\n"
        f"Текст 1 (эталонный): {ground_truth}\n\n"
        f"Текст 2 (ответ модели): {model_answer}\n\n"
        f"Напиши степень соответствия второго текста первому по шкале от 0 до 1, "
        f"где 1 — полностью соответствует по смыслу, 0 — полностью не соответствует по смыслу. "
        f"В ответе укажи только число, округлённое до десятых (например: 0.7)."
    )

    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "user", "content": prompt}
                ],
                timeout=30
            )
            raw_answer = response.choices[0].message.content.strip()

            # Пытаемся распарсить число
            score = float(raw_answer)
            if 0.0 <= score <= 1.0:
                return round(score, 1)
            else:
                print(f"[WARNING] Модель вернула число вне диапазона [0,1]: {score}")
                return None

        except ValueError:
            print(f"[WARNING] Не удалось распарсить ответ модели как число: '{raw_answer}'")
            return None
        except Exception as e:
            print(f"[WARNING] Попытка {attempt} не удалась: {e}")
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                print(f"[ERROR] Не удалось получить оценку после {max_retries} попыток.")
                return None


def evaluate_answers() -> List[Tuple[int, str, str, float]]:
    """
    Оценивает каждый ответ модели относительно эталона.

    Возвращает:
        List[Tuple[int, str, str, float]]: [(номер, эталон, ответ, оценка), ...]
    """
    data = load_data()
    results = []

    for i, (num, ground_truth, model_answer) in enumerate(data, start=1):
        print(f"[INFO] Оценка ответа {num}...")
        score = ask_llm_for_similarity(ground_truth, model_answer)
        if score is None:
            score = 0.0  # или можно оставить NaN, но для Excel лучше 0.0
        results.append((num, ground_truth, model_answer, score))

        # Задержка, чтобы не перегружать API
        time.sleep(API_TIMEOUT)

    return results


def save_evaluation_results(results: List[Tuple[int, str, str, float]], output_file: str = OUTPUT_EVALUATION_FILENAME) -> bool:
    """
    Сохраняет результаты оценки в Excel-файл.

    Столбцы: №, Эталонный ответ, Ответ модели, Оценка соответствия
    """
    try:
        df = pd.DataFrame(results, columns=["№", "Эталонный ответ", "Ответ модели", "Оценка соответствия"])
        df.to_excel(output_file, index=False, engine="openpyxl")

        # Опционально: форматирование (жирный заголовок, автоподбор ширины)
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(output_file)
        ws = wb.active

        # Жирный шрифт для заголовков
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        # Автоподбор ширины
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        print(f"[INFO] Результаты успешно сохранены в {output_file}")
        return True

    except Exception as e:
        print(f"[ERROR] Ошибка при сохранении файла: {e}")
        return False


def main():
    """
    Основная функция программы.
    """
    print("[INFO] Начало оценки соответствия ответов...")
    results = evaluate_answers()
    if results:
        success = save_evaluation_results(results)
        if success:
            print("[INFO] Оценка завершена успешно.")
            # Подсчитаем среднюю оценку
            avg_score = sum(r[3] for r in results) / len(results)
            print(f"[INFO] Средняя оценка соответствия: {avg_score:.2f}")
        else:
            print("[ERROR] Не удалось сохранить результаты.")
    else:
        print("[WARNING] Нет данных для оценки.")


if __name__ == "__main__":
    main()